import datetime
import re

from openpyxl import *

import Main


# this method is highly inefficient and can be replaced by filterAndExtractData
def filterData(fileName, settings):
    wb = load_workbook(fileName)
    pageLoads = wb['Page loads']
    toRemove = []
    startPeriod = settings['period']['startPeriod']
    endPeriod = settings['period']['endPeriod']
    for row in pageLoads.iter_rows(min_row=2):
        if settings['period']['usePeriod']:
            if not (startPeriod <= row[Main.TIMESTAMPINDEX].value <= endPeriod):
                toRemove.append(row[1].row)
                continue

        if settings['students']['whitelist']:
            keep = False
            # students must be on the list to be taken into account
            for requirement in settings['students']['list']:
                if requirement['type'] == "regex":
                    matcher = re.compile(requirement['value'])
                    if matcher.match(row[Main.USERIDINDEX].value):
                        print("user may stay")
                        print(row[Main.USERIDINDEX].value)
                        keep = True
                        break
                elif requirement['value'] == row[Main.USERIDINDEX].value:
                    keep = True
                    print("user may stay")
                    print(row[Main.USERIDINDEX].value)
                    break
            if not keep:
                toRemove.append(row[1].row)

        else:
            for requirement in settings['students']['list']:
                if requirement['type'] == "regex":
                    matcher = re.compile(requirement['value'])
                    if matcher.match(row[Main.USERIDINDEX].value):
                        toRemove.append(row[1].row)
                        break
                elif requirement['value'] == row[Main.USERIDINDEX].value:
                    toRemove.append(row[1].row)
                    break

    toRemove = toRemove[::-1]
    print("done with reading")
    for i in toRemove:
        pageLoads.delete_rows(i)
    print("done removing")
    events = wb['Events']
    toRemove = []
    for row in events.iter_rows(min_row=2):
        if settings['period']['usePeriod']:
            if not (startPeriod <= row[Main.TIMESTAMPINDEX].value <= endPeriod):
                toRemove.append(row[1].row)
                continue

        if settings['students']['whitelist']:
            keep = False
            # students must be on the list to be taken into account
            for requirement in settings['students']['list']:
                if requirement['type'] == "regex":
                    matcher = re.compile(requirement['value'])
                    if matcher.match(row[Main.USERIDINDEX].value):
                        print("user may stay")
                        print(row[Main.USERIDINDEX].value)
                        keep = True
                        break
                elif requirement['value'] == row[Main.USERIDINDEX].value:
                    keep = True
                    print("user may stay")
                    print(row[Main.USERIDINDEX].value)
                    break
            if not keep:
                toRemove.append(row[1].row)

        else:
            for requirement in settings['students']['list']:
                if requirement['type'] == "regex":
                    matcher = re.compile(requirement['value'])
                    if matcher.match(row[Main.USERIDINDEX].value):
                        toRemove.append(row[1].row)
                        break
                elif requirement['value'] == row[Main.USERIDINDEX].value:
                    toRemove.append(row[1].row)
                    break
    print("done reading")
    for i in toRemove:
        events.delete_rows(i)
    wb.save(fileName)

def filterAndExtractData(fileNames, settings,debug=False):
    pageName = "Page loads"
    pageName2 = "Events"
    data = {pageName: [], pageName2: []}
    usePeriod = settings['period']['usePeriod']
    startPeriod = settings['period']['startDate'] if usePeriod else None
    endPeriod = settings['period']['endDate'] if usePeriod else None
    if type(fileNames) is list:
        if debug: print("multpiple files")
        for fileName in fileNames:
            wb = load_workbook(fileName)

            extractAndFilterData(data, endPeriod, pageName, settings, startPeriod, wb, debug)
            extractAndFilterData(data, endPeriod, pageName2, settings, startPeriod, wb, debug)
            wb.close()
            if debug: print("done with" + fileName)
    else:
        wb = load_workbook(fileNames)

        extractAndFilterData(data, endPeriod, pageName, settings, startPeriod, wb, debug)
        extractAndFilterData(data, endPeriod, pageName2, settings, startPeriod, wb, debug)
        wb.close()
    return data


def extractAndFilterData(data, endPeriod, pageName, settings, startPeriod, wb, debug):
    pageLoads = wb[pageName]
    for requirement in settings['students']['list']:
        requirement['value'] = re.compile(requirement['value']) if requirement['type'] == "regex" else requirement['value']

    for row in pageLoads.iter_rows(min_row=2):
        if settings['period']['usePeriod']:
            if not (startPeriod <= row[Main.TIMESTAMPINDEX].value <= endPeriod + datetime.timedelta(days=1)):
                continue  # to the next row

        if settings['students']['whitelist']:  # the given requirements are a whitelist
            for requirement in settings['students']['list']:
                if requirement['type'] == "regex":
                    matcher = requirement['value']
                    if matcher.match(row[Main.USERIDINDEX].value):
                        data[pageName].append(row)
                        break
                elif requirement['value'] == row[Main.USERIDINDEX].value:
                    data[pageName].append(row)
                    break

        else:  # the given requirements are a blacklist
            stay = True
            for requirement in settings['students']['list']:
                if requirement['type'] == "regex":
                    matcher = requirement['value']
                    if matcher.match(row[Main.USERIDINDEX].value):
                        stay = False
                        break
                elif requirement['value'] == row[Main.USERIDINDEX].value:
                    stay = False
                    break
            if stay:
                data[pageName].append(row)
    if debug: print("done with reading " + pageName)
