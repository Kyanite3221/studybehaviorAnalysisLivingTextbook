import sys

from openpyxl import *


def annonymise(fileName):
    wb = load_workbook(fileName)
    pageLoads = wb['Page loads']
    for row in pageLoads.iter_rows(min_row=2):
        row[0].value = str(hash(row[0].value))
    events = wb['Events']
    for row in events.iter_rows(min_row=2):
        row[0].value = str(hash(row[0].value))
    wb.save(fileName)


if __name__ == "--main--":
    if len(sys.argv) < 2:
        sys.argv.append(input("which file would you like to use?\n"))
    annonymise(sys.argv[1])


def annonymiseExtracted(data):
    for row in data["Page loads"]:
        row[0].value = str(hash(row[0].value))
    for row in data["Events"]:
        row[0].value = str(hash(row[0].value))
    return data

def alternateAnonymiseExtracted(data,lookup):
    for row in data["Page loads"]:
        row[0].value = "student " + str(lookup[row[0].value])
    for row in data["Events"]:
        row[0].value = "student " + str(lookup[row[0].value])
    return data
